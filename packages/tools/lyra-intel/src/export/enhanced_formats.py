"""
Enhanced export formats: PDF, SARIF, CSV, Excel.
"""

import json
import csv
from typing import Dict, Any, List, Optional
from datetime import datetime
from io import BytesIO, StringIO

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class PDFExporter:
    """Export analysis results to PDF format."""
    
    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    def export(self, results: Dict[str, Any], output_path: str) -> str:
        """Export results to PDF file."""
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30
        )
        story.append(Paragraph("Lyra Intel Analysis Report", title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Metadata
        metadata = [
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Project", results.get("project_name", "N/A")],
            ["Files Analyzed", str(len(results.get("files", [])))],
            ["Total Issues", str(len(results.get("issues", [])))]
        ]
        
        metadata_table = Table(metadata, colWidths=[2*inch, 4*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Summary metrics
        story.append(Paragraph("Summary", styles['Heading2']))
        metrics = results.get("metrics", {})
        
        summary_data = [
            ["Metric", "Value"],
            ["Complexity", f"{metrics.get('complexity', 0):.2f}"],
            ["Maintainability", f"{metrics.get('maintainability', 0):.2f}"],
            ["Test Coverage", f"{metrics.get('coverage', 0):.1f}%"],
            ["Security Score", f"{metrics.get('security_score', 0):.1f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(PageBreak())
        
        # Issues
        story.append(Paragraph("Issues Found", styles['Heading2']))
        issues = results.get("issues", [])
        
        for issue in issues[:50]:  # Limit to first 50 issues
            issue_text = f"""
            <b>Severity:</b> {issue.get('severity', 'Unknown')}<br/>
            <b>Type:</b> {issue.get('type', 'Unknown')}<br/>
            <b>File:</b> {issue.get('file', 'Unknown')}<br/>
            <b>Line:</b> {issue.get('line', 'N/A')}<br/>
            <b>Message:</b> {issue.get('message', 'No description')}
            """
            story.append(Paragraph(issue_text, styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
        
        doc.build(story)
        return output_path


class SARIFExporter:
    """Export security findings to SARIF format."""
    
    def export(self, results: Dict[str, Any], output_path: str) -> str:
        """Export results to SARIF file."""
        sarif = {
            "$schema": "https://raw.githubusercontent.com/oasis-tcs/sarif-spec/master/Schemata/sarif-schema-2.1.0.json",
            "version": "2.1.0",
            "runs": [
                {
                    "tool": {
                        "driver": {
                            "name": "Lyra Intel",
                            "version": "1.0.0",
                            "informationUri": "https://github.com/nirholas/lyra-intel",
                            "rules": self._generate_rules(results)
                        }
                    },
                    "results": self._generate_results(results)
                }
            ]
        }
        
        with open(output_path, 'w') as f:
            json.dump(sarif, f, indent=2)
        
        return output_path
    
    def _generate_rules(self, results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate SARIF rules from issue types."""
        rules = []
        issue_types = set()
        
        for issue in results.get("issues", []):
            issue_type = issue.get("type", "Unknown")
            if issue_type not in issue_types:
                issue_types.add(issue_type)
                rules.append({
                    "id": issue_type,
                    "name": issue_type,
                    "shortDescription": {
                        "text": issue.get("message", "")
                    },
                    "fullDescription": {
                        "text": issue.get("description", issue.get("message", ""))
                    },
                    "helpUri": f"https://github.com/nirholas/lyra-intel/docs/rules/{issue_type}"
                })
        
        return rules
    
    def _generate_results(self, results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate SARIF results from issues."""
        sarif_results = []
        
        for issue in results.get("issues", []):
            result = {
                "ruleId": issue.get("type", "Unknown"),
                "level": self._map_severity(issue.get("severity", "warning")),
                "message": {
                    "text": issue.get("message", "No description")
                },
                "locations": [
                    {
                        "physicalLocation": {
                            "artifactLocation": {
                                "uri": issue.get("file", "unknown")
                            },
                            "region": {
                                "startLine": issue.get("line", 1),
                                "startColumn": issue.get("column", 1)
                            }
                        }
                    }
                ]
            }
            
            if "fixes" in issue:
                result["fixes"] = issue["fixes"]
            
            sarif_results.append(result)
        
        return sarif_results
    
    def _map_severity(self, severity: str) -> str:
        """Map issue severity to SARIF level."""
        mapping = {
            "critical": "error",
            "high": "error",
            "medium": "warning",
            "low": "note",
            "info": "note"
        }
        return mapping.get(severity.lower(), "warning")


class CSVExporter:
    """Export analysis results to CSV format."""
    
    def export(self, results: Dict[str, Any], output_path: str) -> str:
        """Export results to CSV file."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow([
                "File",
                "Line",
                "Column",
                "Severity",
                "Type",
                "Message",
                "Rule"
            ])
            
            # Write issues
            for issue in results.get("issues", []):
                writer.writerow([
                    issue.get("file", ""),
                    issue.get("line", ""),
                    issue.get("column", ""),
                    issue.get("severity", ""),
                    issue.get("type", ""),
                    issue.get("message", ""),
                    issue.get("rule", "")
                ])
        
        return output_path
    
    def export_metrics(self, results: Dict[str, Any], output_path: str) -> str:
        """Export metrics to CSV file."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(["Metric", "Value"])
            
            # Write metrics
            metrics = results.get("metrics", {})
            for key, value in metrics.items():
                writer.writerow([key, value])
        
        return output_path


class ExcelExporter:
    """Export analysis results to Excel format."""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def export(self, results: Dict[str, Any], output_path: str) -> str:
        """Export results to Excel file."""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, results)
        self._create_issues_sheet(wb, results)
        self._create_metrics_sheet(wb, results)
        self._create_files_sheet(wb, results)
        
        wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, results: Dict[str, Any]):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Lyra Intel Analysis Report"
        ws['A1'].font = Font(size=18, bold=True)
        
        # Metadata
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Project:"
        ws['B4'] = results.get("project_name", "N/A")
        
        ws['A5'] = "Files Analyzed:"
        ws['B5'] = len(results.get("files", []))
        
        ws['A6'] = "Total Issues:"
        ws['B6'] = len(results.get("issues", []))
        
        # Format
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_issues_sheet(self, wb: openpyxl.Workbook, results: Dict[str, Any]):
        """Create issues sheet."""
        ws = wb.create_sheet("Issues")
        
        # Header
        headers = ["File", "Line", "Column", "Severity", "Type", "Message"]
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for issue in results.get("issues", []):
            ws.append([
                issue.get("file", ""),
                issue.get("line", ""),
                issue.get("column", ""),
                issue.get("severity", ""),
                issue.get("type", ""),
                issue.get("message", "")
            ])
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_metrics_sheet(self, wb: openpyxl.Workbook, results: Dict[str, Any]):
        """Create metrics sheet."""
        ws = wb.create_sheet("Metrics")
        
        # Header
        ws.append(["Metric", "Value"])
        
        # Format header
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Data
        metrics = results.get("metrics", {})
        for key, value in metrics.items():
            ws.append([key, value])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_files_sheet(self, wb: openpyxl.Workbook, results: Dict[str, Any]):
        """Create files sheet."""
        ws = wb.create_sheet("Files")
        
        # Header
        ws.append(["File", "Language", "Lines", "Complexity"])
        
        # Format header
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)
        ws['D1'].font = Font(bold=True)
        
        # Data
        for file_info in results.get("files", []):
            ws.append([
                file_info.get("path", ""),
                file_info.get("language", ""),
                file_info.get("lines", 0),
                file_info.get("complexity", 0)
            ])
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
